import asyncio
import re
import html
import sqlite3
import os
from datetime import datetime, timezone
from aiogram import Bot, Dispatcher, types, F
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram.webhook.aiohttp_server import SimpleRequestHandler, setup_application
from aiohttp import web, MultipartReader
import logging
import io

# Silence noisy logs
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# =========================================================
# 🔐 HARD-CODED CONFIG — REPLACE THESE VALUES
# =========================================================
BOT_TOKEN = "8119637737:AAEPDFrHrUeAHWHSHkR7ZSE_Hp1vg3sYs9c"
ADMIN_ID = 6812877108          # Your Telegram user ID
GROUP_ID = -1003021667823     # Your IVASMS forward group ID
WEBHOOK_URL = "https://otp-xfm2.onrender.com/webhook"  # Update after deploy
PORT = 10000

OWNER_LINK = "https://t.me/cryptoearn36"
NUMBERS_CHANNEL = "https://t.me/oxfreeOTP"

# Earnings per SMS
EARNINGS_PER_SMS = 1.0000  # ৳1.0000 per OTP

# Minimum withdrawal
MIN_WITHDRAWAL = 250.0

# Database path (Render ephemeral disk)
DB_PATH = "/tmp/ivasms_bot.db"

# =========================================================
# COUNTRY & SERVICE MAPPINGS (from your Excel data)
# =========================================================
COUNTRY_FLAGS = {
    "998": "🇺🇿 Uzbekistan", "225": "🇨🇮 Ivory Coast", "93": "🇦🇫 Afghanistan",
    "234": "🇳🇬 Nigeria", "967": "🇾🇪 Yemen", "63": "🇵🇭 Philippines",
    "20": "🇪🇬 Egypt", "994": "🇦🇿 Azerbaijan", "996": "🇰🇬 Kyrgyzstan",
    "228": "🇹🇬 Togo", "961": "🇱🇧 Lebanon", "229": "🇧🇯 Benin",
    "977": "🇳🇵 Nepal", "963": "🇸🇾 Syria", "216": "🇹🇳 Tunisia"
}

SERVICES = {
    "whatsapp": "WhatsApp", "facebook": "Facebook", "telegram": "Telegram",
    "google": "Google", "instagram": "Instagram", "tiktok": "TikTok",
    "apple": "Apple", "1xbet": "1xBet", "melbet": "Melbet", "exness": "Exness",
    "wildberries": "Wildberries", "betwinner": "Betwinner", "netflix": "Netflix",
    "microsoft": "Microsoft", "binance": "Binance", "premierbet": "PremierBet"
}

# =========================================================
# NUMBERS FROM YOUR EXCEL SHEET (My Numbers Sheet (2).xlsx)
# =========================================================
LOCAL_NUMBERS = [
    {"country": "UZBEKISTAN", "number": "998992842855"},
    {"country": "UZBEKISTAN", "number": "998992849932"},
    {"country": "UZBEKISTAN", "number": "998992847741"},
    {"country": "UZBEKISTAN", "number": "998992847688"},
    {"country": "UZBEKISTAN", "number": "998992848697"},
    {"country": "UZBEKISTAN", "number": "998992843332"},
    {"country": "UZBEKISTAN", "number": "998992848571"},
    {"country": "UZBEKISTAN", "number": "998992849824"},
    {"country": "UZBEKISTAN", "number": "998992841083"},
    {"country": "UZBEKISTAN", "number": "998992841176"},
    {"country": "UZBEKISTAN", "number": "998992840698"},
    {"country": "UZBEKISTAN", "number": "998992849662"},
    {"country": "UZBEKISTAN", "number": "998992841583"},
    {"country": "UZBEKISTAN", "number": "998992841110"},
    {"country": "UZBEKISTAN", "number": "998992844753"},
    {"country": "UZBEKISTAN", "number": "998992841407"},
    {"country": "UZBEKISTAN", "number": "998992841677"},
    {"country": "UZBEKISTAN", "number": "998992846699"},
    {"country": "UZBEKISTAN", "number": "998992847638"},
    {"country": "UZBEKISTAN", "number": "998992845430"},
    {"country": "UZBEKISTAN", "number": "998992843833"},
    {"country": "UZBEKISTAN", "number": "998992848831"},
    {"country": "UZBEKISTAN", "number": "998992843431"},
    {"country": "UZBEKISTAN", "number": "998992846972"},
    {"country": "UZBEKISTAN", "number": "998992840149"},
    {"country": "UZBEKISTAN", "number": "998992849722"},
    {"country": "UZBEKISTAN", "number": "998992849344"},
    {"country": "UZBEKISTAN", "number": "998992844341"},
    {"country": "UZBEKISTAN", "number": "998992841460"},
    {"country": "UZBEKISTAN", "number": "998992846541"},
    {"country": "UZBEKISTAN", "number": "998992842444"},
    {"country": "UZBEKISTAN", "number": "998992846733"},
    {"country": "UZBEKISTAN", "number": "998992842176"},
    {"country": "UZBEKISTAN", "number": "998992844291"},
    {"country": "UZBEKISTAN", "number": "998992849878"},
    {"country": "UZBEKISTAN", "number": "998992843545"},
    {"country": "UZBEKISTAN", "number": "998992849865"},
    {"country": "UZBEKISTAN", "number": "998992842054"},
    {"country": "UZBEKISTAN", "number": "998992842202"},
    {"country": "UZBEKISTAN", "number": "998992847996"},
    {"country": "UZBEKISTAN", "number": "998992847885"},
    {"country": "UZBEKISTAN", "number": "998992840169"},
    {"country": "UZBEKISTAN", "number": "998992848047"},
    {"country": "UZBEKISTAN", "number": "998992844434"},
    {"country": "UZBEKISTAN", "number": "998992842057"},
    {"country": "UZBEKISTAN", "number": "998992847279"},
    {"country": "UZBEKISTAN", "number": "998992846400"},
    {"country": "UZBEKISTAN", "number": "998992844387"},
    {"country": "UZBEKISTAN", "number": "998992842654"},
    {"country": "UZBEKISTAN", "number": "998992845138"},
    {"country": "UZBEKISTAN", "number": "998992847862"},
    {"country": "UZBEKISTAN", "number": "998992847090"},
    {"country": "UZBEKISTAN", "number": "998992845141"},
    {"country": "UZBEKISTAN", "number": "998992849726"},
    {"country": "UZBEKISTAN", "number": "998992847569"},
    {"country": "UZBEKISTAN", "number": "998992846696"},
    {"country": "UZBEKISTAN", "number": "998992840342"},
    {"country": "UZBEKISTAN", "number": "998992845261"},
    {"country": "UZBEKISTAN", "number": "998992845433"},
    {"country": "UZBEKISTAN", "number": "998992840009"},
    {"country": "UZBEKISTAN", "number": "998992849719"},
    {"country": "UZBEKISTAN", "number": "998992847842"},
    {"country": "UZBEKISTAN", "number": "998992840011"},
    {"country": "UZBEKISTAN", "number": "998992843462"},
    {"country": "UZBEKISTAN", "number": "998992842139"},
    {"country": "UZBEKISTAN", "number": "998992842388"},
    {"country": "UZBEKISTAN", "number": "998992843174"},
    {"country": "UZBEKISTAN", "number": "998992848675"},
    {"country": "UZBEKISTAN", "number": "998992846589"},
    {"country": "UZBEKISTAN", "number": "998992840958"},
    {"country": "UZBEKISTAN", "number": "998992844277"},
    {"country": "UZBEKISTAN", "number": "998992846276"},
    {"country": "UZBEKISTAN", "number": "998992843999"},
    {"country": "UZBEKISTAN", "number": "998992849236"},
    {"country": "UZBEKISTAN", "number": "998992841053"},
    {"country": "UZBEKISTAN", "number": "998992840950"},
    {"country": "UZBEKISTAN", "number": "998992841049"},
    {"country": "UZBEKISTAN", "number": "998992846577"},
    {"country": "UZBEKISTAN", "number": "998992844929"},
    {"country": "UZBEKISTAN", "number": "998992847317"},
    {"country": "UZBEKISTAN", "number": "998992842314"},
    {"country": "UZBEKISTAN", "number": "998992840291"},
    {"country": "UZBEKISTAN", "number": "998992847145"},
    {"country": "UZBEKISTAN", "number": "998992846194"},
    {"country": "UZBEKISTAN", "number": "998992843397"},
    {"country": "UZBEKISTAN", "number": "998992843455"},
    {"country": "UZBEKISTAN", "number": "998992847099"},
    {"country": "UZBEKISTAN", "number": "998992845577"},
    {"country": "UZBEKISTAN", "number": "998992844958"},
    {"country": "UZBEKISTAN", "number": "998992842652"},
    {"country": "UZBEKISTAN", "number": "998992843931"},
    {"country": "UZBEKISTAN", "number": "998992842509"},
    {"country": "UZBEKISTAN", "number": "998992848490"},
    {"country": "UZBEKISTAN", "number": "998992842758"},
    {"country": "UZBEKISTAN", "number": "998992842371"},
    {"country": "UZBEKISTAN", "number": "998992848726"},
    {"country": "UZBEKISTAN", "number": "998992846296"},
    {"country": "UZBEKISTAN", "number": "998992845856"},
    {"country": "UZBEKISTAN", "number": "998992846126"},
    {"country": "UZBEKISTAN", "number": "998992844245"}
]

# =========================================================
# DATABASE SETUP
# =========================================================
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        user_id INTEGER PRIMARY KEY,
        username TEXT,
        numbers TEXT DEFAULT '[]',
        balance REAL DEFAULT 0.0
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS withdrawals (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        amount REAL,
        bkash TEXT,
        status TEXT DEFAULT 'pending'
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS otps (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        number TEXT,
        otp TEXT,
        full_msg TEXT,
        service TEXT,
        country TEXT,
        fetched_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS available_numbers (
        number TEXT PRIMARY KEY,
        country TEXT,
        range_info TEXT,
        assigned_to INTEGER DEFAULT NULL
    )''')
    # Populate available numbers from Excel
    for n in LOCAL_NUMBERS:
        c.execute("INSERT OR IGNORE INTO available_numbers (number, country, range_info) VALUES (?, ?, ?)",
                  (n["number"], n["country"], f"{n['country']} Range"))
    conn.commit()
    conn.close()

def get_connection():
    return sqlite3.connect(DB_PATH)

# =========================================================
# UTILS
# =========================================================
def detect_country(number):
    s = number.lstrip("+")
    for prefix, name in COUNTRY_FLAGS.items():
        if s.startswith(prefix):
            return name
    return "🌍 Unknown"

def detect_service(text):
    t = text.lower()
    for k in sorted(SERVICES, key=len, reverse=True):
        if k in t:
            return SERVICES[k]
    return "Service"

def extract_otps(text):
    if m := re.search(r"(?:code|is|:)\s*(\b\d{4,8}\b)", text, re.IGNORECASE):
        return [m.group(1)]
    return re.findall(r"\b(\d{4,8})\b", text)

def mask_number(num):
    s = num.strip()
    if len(s) <= 10:
        return s
    return s[:7] + "****" + s[-3:]

def otp_exists(number, otp):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT 1 FROM otps WHERE number = ? AND otp = ?", (number, otp))
    return c.fetchone() is not None

def save_otp(number, otp, msg, service, country):
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO otps (number, otp, full_msg, service, country, fetched_at) VALUES (?, ?, ?, ?, ?, ?)",
              (number, otp, msg, service, country, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

def get_user_by_number(number):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT user_id FROM users WHERE numbers LIKE ?", (f'%"{number}"%',))
    row = c.fetchone()
    conn.close()
    return row[0] if row else None

def credit_user(user_id, amount=EARNINGS_PER_SMS):
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE users SET balance = balance + ? WHERE user_id = ?", (amount, user_id))
    conn.commit()
    conn.close()

def get_available_countries():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT country, COUNT(*) FROM available_numbers WHERE assigned_to IS NULL GROUP BY country")
    rows = c.fetchall()
    conn.close()
    return dict(rows)

def assign_number(user_id, country):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT number FROM available_numbers WHERE country = ? AND assigned_to IS NULL LIMIT 1", (country,))
    row = c.fetchone()
    if not row:
        return None
    number = row[0]
    c.execute("UPDATE available_numbers SET assigned_to = ? WHERE number = ?", (user_id, number))
    c.execute("SELECT numbers FROM users WHERE user_id = ?", (user_id,))
    user_row = c.fetchone()
    if user_row:
        numbers = eval(user_row[0])
        numbers.append(number)
        c.execute("UPDATE users SET numbers = ? WHERE user_id = ?", (str(numbers), user_id))
    else:
        c.execute("INSERT INTO users (user_id, numbers) VALUES (?, ?)", (user_id, str([number])))
    conn.commit()
    conn.close()
    return number

def add_numbers_from_excel(rows):
    conn = get_connection()
    c = conn.cursor()
    added = 0
    for row in rows:
        if len(row) < 2:
            continue
        number = str(row[1]).strip()
        if not re.match(r"\+?\d{6,15}", number):
            continue
        country = detect_country(number)
        range_info = str(row[0]) if row[0] else "Uploaded"
        c.execute("INSERT OR IGNORE INTO available_numbers (number, country, range_info) VALUES (?, ?, ?)", (number, country, range_info))
        added += 1
    conn.commit()
    conn.close()
    return added

# =========================================================
# TELEGRAM BOT
# =========================================================
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()
user_states = {}

# OTP handler (from IVASMS group)
@dp.message(F.chat.id == GROUP_ID)
async def handle_group_message(m: types.Message):
    if not m.text:
        return
    lines = m.text.split('\n')
    try:
        number_line = next(l for l in lines if 'Number:' in l)
        number = number_line.split('Number:')[1].strip().split()[0]
        otp_line = next(l for l in lines if 'OTP Code:' in l)
        otp = otp_line.split('OTP Code:')[1].strip()
    except:
        return
    
    if otp_exists(number, otp):
        return
    
    service = detect_service(m.text)
    country = detect_country(number)
    entry = {"number": number, "otp": otp, "full_msg": m.text, "service": service, "country": country}
    save_otp(number, otp, m.text, service, country)
    
    # Forward to group with branding
    text = (
        f"🔔 <b>NEW OTP DETECTED</b>\n🆕\n\n"
        f"🕰 <b>Time:</b> {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"🌍 <b>Country:</b> {country}\n"
        f"⚙️ <b>Service:</b> {service}\n"
        f"☎️ <b>Number:</b> {mask_number(number)}\n"
        f"🔑 <b>OTP:</b> <code>{otp}</code>\n\n"
        f"📩 <b>Full Message:</b>\n"
        f"<pre>{html.escape(m.text)}</pre>"
    )
    kb = types.InlineKeyboardMarkup(inline_keyboard=[
        [types.InlineKeyboardButton("👑 ×°𝓞𝔀𝓷𝓮𝓻°× 👑", url=OWNER_LINK),
         types.InlineKeyboardButton("༄ 𝐃𝐞𝐯𝐞𝐥𝐨𝐩𝐞𝐫 𒆜", url="https://t.me/BashOnChain ")],
        [types.InlineKeyboardButton("★彡[ᴀʟʟ ɴᴜᴍʙᴇʀꜱ]彡★", url=NUMBERS_CHANNEL)]
    ])
    await bot.send_message(GROUP_ID, text, reply_markup=kb)
    
    # Send to assigned user
    user_id = get_user_by_number(number)
    if user_id:
        credit_user(user_id)
        await bot.send_message(
            user_id,
            f"💰 New OTP!\n🔑 <code>{otp}</code>\n📞 <code>{number}</code>\n\n✅ ৳{EARNINGS_PER_SMS:.4f} added!",
            parse_mode=ParseMode.HTML
        )

# User commands
@dp.message(F.text == "/start")
async def start(m: types.Message):
    kb = types.InlineKeyboardMarkup(inline_keyboard=[
        [types.InlineKeyboardButton("🎁 Get Number", callback_data='get_number')],
        [types.InlineKeyboardButton("👤 Account", callback_data='account')],
        [types.InlineKeyboardButton("💰 Withdraw", callback_data='withdraw')]
    ])
    await m.answer(
        "👋 Welcome to IVASMS OTP Bot!\n\n"
        f"💰 Earn ৳{EARNINGS_PER_SMS:.4f} per OTP\n"
        "📱 Get a number → Use it → Get paid!",
        reply_markup=kb
    )

@dp.callback_query(F.data == "get_number")
async def get_number(q: types.CallbackQuery):
    countries = get_available_countries()
    if not countries:
        await q.message.edit_text("❌ No numbers available.")
        return
    kb = [[types.InlineKeyboardButton(f"{k} ({v})", callback_data=f'country_{k}')] for k, v in countries.items()]
    kb.append([types.InlineKeyboardButton("❌ Cancel", callback_data='cancel')])
    await q.message.edit_text("🌍 Select country:", reply_markup=types.InlineKeyboardMarkup(inline_keyboard=kb))

@dp.callback_query(F.data.startswith("country_"))
async def select_country(q: types.CallbackQuery):
    country = q.data.split('_', 1)[1]
    number = assign_number(q.from_user.id, country)
    if number:
        await q.message.edit_text(f"✅ Assigned:\n📞 <code>{number}</code>\n🌍 {country}", parse_mode=ParseMode.HTML)
    else:
        await q.message.edit_text("❌ No available numbers.")

@dp.callback_query(F.data == "account")
async def account(q: types.CallbackQuery):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT balance, numbers FROM users WHERE user_id = ?", (q.from_user.id,))
    row = c.fetchone()
    balance = row[0] if row else 0.0
    numbers = eval(row[1]) if row else []
    conn.close()
    nums = "\n".join([f"• <code>{n}</code>" for n in numbers]) if numbers else "None"
    await q.message.edit_text(
        f"👤 Your Account\n\n"
        f"💰 Balance: ৳{balance:.2f}\n"
        f"📱 Assigned Numbers:\n{nums}",
        parse_mode=ParseMode.HTML
    )

@dp.callback_query(F.data == "withdraw")
async def withdraw(q: types.CallbackQuery):
    await q.message.edit_text("📲 Send your Bkash number (11 digits, e.g., 017XXXXXXXX):")
    user_states[q.from_user.id] = "awaiting_bkash"

@dp.message(F.text)
async def handle_msg(m: types.Message):
    if user_states.get(m.from_user.id) == "awaiting_bkash":
        bkash = m.text.strip()
        if not (bkash.startswith('01') and len(bkash) == 11 and bkash.isdigit()):
            await m.answer("❌ Invalid Bkash number. Must be 11 digits starting with 01.")
            return
        conn = get_connection()
        c = conn.cursor()
        c.execute("SELECT balance FROM users WHERE user_id = ?", (m.from_user.id,))
        row = c.fetchone()
        balance = row[0] if row else 0.0
        if balance < MIN_WITHDRAWAL:
            await m.answer(f"❌ Minimum withdrawal is ৳{MIN_WITHDRAWAL:.2f}")
            return
        c.execute("INSERT INTO withdrawals (user_id, amount, bkash) VALUES (?, ?, ?)", (m.from_user.id, balance, bkash))
        c.execute("UPDATE users SET balance = 0 WHERE user_id = ?", (m.from_user.id,))
        conn.commit()
        conn.close()
        await m.answer(
            f"✅ Withdrawal request submitted!\n\n"
            f"Amount: ৳{balance:.2f}\n"
            f"Bkash: {bkash}\n\n"
            "Admin will process your request soon."
        )
        del user_states[m.from_user.id]

# Admin panel
@dp.message(F.text == "/admin")
async def admin(m: types.Message):
    if m.from_user.id != ADMIN_ID:
        return
    kb = types.InlineKeyboardMarkup(inline_keyboard=[
        [types.InlineKeyboardButton("📤 Upload Excel", callback_data='admin_upload')],
        [types.InlineKeyboardButton("💸 Withdrawals", callback_data='admin_withdrawals')],
        [types.InlineKeyboardButton("👥 Users", callback_data='admin_users')]
    ])
    await m.answer("🔐 Admin Panel", reply_markup=kb)

@dp.callback_query(F.data == "admin_upload")
async def admin_upload(q: types.CallbackQuery):
    await q.message.edit_text("📤 Send an Excel file (.xlsx) with columns: Range, Number")
    user_states[ADMIN_ID] = "awaiting_excel"

@dp.message(F.document)
async def handle_excel(m: types.Message):
    if m.from_user.id != ADMIN_ID or user_states.get(ADMIN_ID) != "awaiting_excel":
        return
    if not m.document.file_name.endswith('.xlsx'):
        await m.answer("❌ Please send a .xlsx file.")
        return
    try:
        file = await bot.get_file(m.document.file_id)
        file_bytes = await bot.download_file(file.file_path)
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(file_bytes.read()))
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                rows.append(row)
        count = add_numbers_from_excel(rows)
        await m.answer(f"✅ Added {count} numbers from Excel!")
    except Exception as e:
        await m.answer(f"❌ Failed to parse Excel: {e}")
    finally:
        del user_states[ADMIN_ID]

@dp.callback_query(F.data == "admin_withdrawals")
async def admin_wd(q: types.CallbackQuery):
    if q.from_user.id != ADMIN_ID:
        return
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, user_id, amount, bkash FROM withdrawals WHERE status = 'pending'")
    pending = c.fetchall()
    conn.close()
    if not pending:
        await q.message.edit_text("✅ No pending withdrawals.")
        return
    text = "💸 Pending Withdrawals:\n\n" + "\n".join([f"ID: {w[0]} | User: {w[1]} | ৳{w[2]:.2f} | {w[3]}" for w in pending])
    text += f"\n\nUse /approve <id> to approve."
    await q.message.edit_text(text)

@dp.callback_query(F.data == "admin_users")
async def admin_users(q: types.CallbackQuery):
    if q.from_user.id != ADMIN_ID:
        return
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT user_id, username, balance, numbers FROM users")
    users = c.fetchall()
    conn.close()
    if not users:
        await q.message.edit_text("📭 No users found.")
        return
    text = "👥 Users:\n\n"
    for uid, uname, bal, nums in users:
        count = len(eval(nums))
        name = uname or f"User{uid}"
        text += f"ID: <code>{uid}</code> | {name} | Balance: ৳{bal:.2f} | Numbers: {count}\n"
    await q.message.edit_text(text, parse_mode=ParseMode.HTML)

@dp.message(F.text.startswith("/approve"))
async def approve(m: types.Message):
    if m.from_user.id != ADMIN_ID:
        return
    try:
        wid = int(m.text.split()[1])
        conn = get_connection()
        c = conn.cursor()
        c.execute("UPDATE withdrawals SET status = 'paid' WHERE id = ?", (wid,))
        conn.commit()
        conn.close()
        await m.answer(f"✅ Withdrawal #{wid} approved and paid!")
    except:
        await m.answer("Usage: /approve <id>")

# =========================================================
# RENDER HEALTH CHECK
# =========================================================
async def healthz(request):
    return web.Response(text="OK", content_type="text/plain")

# =========================================================
# MAIN
# =========================================================
async def on_startup(app):
    init_db()
    await bot.set_webhook(WEBHOOK_URL)
    logger.info("✅ Bot started on Render with webhook.")

async def on_shutdown(app):
    await bot.delete_webhook(drop_pending_updates=True)

if __name__ == "__main__":
    app = web.Application()
    app.router.add_get("/healthz", healthz)
    webhook_requests_handler = SimpleRequestHandler(dispatcher=dp, bot=bot)
    webhook_requests_handler.register(app, path="/webhook")
    app.on_startup.append(on_startup)
    app.on_shutdown.append(on_shutdown)
    web.run_app(app, host="0.0.0.0", port=PORT)