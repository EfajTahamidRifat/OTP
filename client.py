# bot.py
"""
Termux-compatible Telegram bot for IVASMS OTP-forwarding & rewards.
Make sure to set the CONFIG section before running.
"""

import os
import re
import json
import time
import uuid
import hashlib
import asyncio
from io import BytesIO
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
)
from telegram.ext import (
    ApplicationBuilder, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, filters
)

# ---------------------------
# CONFIG — FILL THESE VALUES
# ---------------------------
TELEGRAM_TOKEN = "YOUR_BOT_TOKEN"
ADMIN_ID = 6812877108
TEAM_CHAT_ID = -1002224177347

IVASMS_EMAIL = "your_ivasms_email@example.com"
IVASMS_PASSWORD = "your_ivasms_password"
IVASMS_URL = "https://www.ivasms.com/portal/live/test_sms"  # inbox page

POLL_SECONDS = 25           # how often to poll IVASMS
REWARD_PER_OTP = 1          # Tk per OTP
MIN_WITHDRAW = 250          # minimum Tk to request withdrawal

# Filenames (in the same folder)
USERS_FILE = "users.json"
NUMBERS_FILE = "numbers.json"
SEEN_FILE = "seen.json"
WITHD_FILE = "withdrawals.json"

# ---------------------------
# Ensure data files exist
# ---------------------------
def ensure_file(path, default):
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(default, f, ensure_ascii=False, indent=2)

ensure_file(USERS_FILE, {})
ensure_file(NUMBERS_FILE, [])
ensure_file(SEEN_FILE, [])
ensure_file(WITHD_FILE, [])

# ---------------------------
# Utilities
# ---------------------------
def load_json(path, default=None):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default if default is not None else {}

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def normalize_num(n):
    return re.sub(r"[^\d]", "", (n or ""))

def hash_msg(sender, content):
    return hashlib.sha1((sender + "|" + content).encode()).hexdigest()

# ---------------------------
# IVASMS login & fetch (blocking)
# ---------------------------
def create_logged_session():
    s = requests.Session()
    try:
        r = s.get("https://www.ivasms.com/portal/login", timeout=20)
    except Exception as e:
        print("Failed to fetch login page:", e)
        return None

    data = {}
    soup = BeautifulSoup(r.text, "html.parser")
    for inp in soup.find_all("input", {"type": "hidden"}):
        name = inp.get("name")
        val = inp.get("value", "") or ""
        if name:
            data[name] = val

    data.update({
        "email": IVASMS_EMAIL or "",
        "password": IVASMS_PASSWORD or ""
    })

    try:
        resp = s.post("https://www.ivasms.com/portal/login", data=data, timeout=20)
        if resp.status_code not in (200, 302):
            print("Login POST status:", resp.status_code)
        return s
    except Exception as e:
        print("Login failed:", e)
        return None

def fetch_messages_from_page(session):
    """
    Returns list of dicts: {id, sender_or_to, content}
    This is a generic parser — adjust if IVASMS HTML is different.
    """
    try:
        r = session.get(IVASMS_URL, timeout=20)
    except Exception as e:
        print("Failed to load IVASMS inbox:", e)
        return []

    if r.status_code != 200:
        print("IVASMS page status:", r.status_code)
        return []

    soup = BeautifulSoup(r.text, "html.parser")
    rows = soup.select("table tr")
    results = []
    for row in rows:
        cells = row.find_all(("td","th"))
        if not cells or len(cells) < 2:
            continue
        sender_or_to = cells[0].get_text(strip=True)
        content = cells[1].get_text(strip=True)
        if len(cells) >= 3:
            extra = cells[2].get_text(strip=True)
            if extra:
                content = f"{content} {extra}"
        uid = hash_msg(sender_or_to, content)
        results.append({"id": uid, "sender_or_to": sender_or_to, "content": content})
    return results

# ---------------------------
# User & Number Management
# ---------------------------
def load_users():
    return load_json(USERS_FILE, {})

def save_users(u):
    save_json(USERS_FILE, u)

def load_numbers():
    return load_json(NUMBERS_FILE, [])

def save_numbers(n):
    save_json(NUMBERS_FILE, n)

def load_seen():
    return set(load_json(SEEN_FILE, []))

def save_seen(s):
    save_json(SEEN_FILE, list(s))

def load_withdrawals():
    return load_json(WITHD_FILE, [])

def save_withdrawals(w):
    save_json(WITHD_FILE, w)

def find_free_number(numbers):
    for n in numbers:
        if not n.get("assigned_to"):
            return n
    return None

def assign_number_to_user(numbers, users, number_entry, tg_id, username):
    number_entry["assigned_to"] = str(tg_id)
    users.setdefault(str(tg_id), {})
    users[str(tg_id)].update({
        "username": username or "",
        "number": number_entry["number"],
        "balance": users.get(str(tg_id), {}).get("balance", 0)
    })
    save_numbers(numbers)
    save_users(users)
    return users[str(tg_id)]

def get_user_by_number(users, number):
    for tg_id, u in users.items():
        if u.get("number") == number:
            return tg_id, u
    return None, None

def find_target_for_sms(users, numbers, sender_or_to, content):
    norm_text = normalize_num(sender_or_to + " " + content)
    # check numbers list
    for n in numbers:
        num = n.get("number")
        if not num:
            continue
        if normalize_num(num) and normalize_num(num) in norm_text:
            tg_id, u = get_user_by_number(users, num)
            if tg_id:
                return tg_id, u, num
    # fallback check user assigned numbers
    for tg_id, u in users.items():
        num = u.get("number")
        if not num:
            continue
        if normalize_num(num) in norm_text:
            return tg_id, u, num
    return None, None, None

# ---------------------------
# OTP parsing helpers
# ---------------------------
OTP_REGEXES = [
    r"\b(\d{4,8})\b",              # 4-8 digit codes
    r"code[:\s]*([0-9]{4,8})",
    r"OTP[:\s]*([0-9]{4,8})",
    r"verification\s*code[:\s]*([0-9]{4,8})",
]

def extract_otp(text):
    # Try to find the most likely OTP number (longest 4-8 digit sequence)
    found = []
    for regex in OTP_REGEXES:
        m = re.findall(regex, text, flags=re.IGNORECASE)
        if m:
            found.extend(m)
    if not found:
        # fallback: pick first 4-6 digit substring
        m = re.findall(r"\b(\d{4,6})\b", text)
        if m:
            found = m
    # choose the longest candidate
    if not found:
        return ""
    candidate = max(found, key=len)
    return candidate

def detect_provider_and_country(text):
    # Best-effort keywords — refine with real IVASMS table if possible
    provider = ""
    country = ""
    # provider keywords
    p_keys = ["Melbet","Facebook","Google","WhatsApp","Twitter","Instagram","PayPal","Bkash"]
    for p in p_keys:
        if p.lower() in text.lower():
            provider = p
            break
    # country keywords or dialing codes
    country_map = {
        "225": "IVORY COAST",
        "880": "BANGLADESH",
        "91": "INDIA",
        "1": "USA",
        "44": "UK",
    }
    # look for known dialcodes in text
    for code, name in country_map.items():
        if code in normalize_num(text):
            country = name
            break
    return provider, country

# ---------------------------
# Telegram send wrapper
# ---------------------------
async def send_message(context: ContextTypes.DEFAULT_TYPE, chat_id, text, parse_mode="Markdown"):
    try:
        await context.bot.send_message(chat_id=chat_id, text=text, parse_mode=parse_mode)
        return True
    except Exception as e:
        print("Telegram send error:", e)
        return False

# ---------------------------
# Telegram UI and Handlers
# ---------------------------
MAIN_MENU = InlineKeyboardMarkup([
    [InlineKeyboardButton("📱 Get Number", callback_data="get_number")],
    [InlineKeyboardButton("💰 My Balance", callback_data="my_balance"),
     InlineKeyboardButton("📨 My OTPs", callback_data="my_otps")],
    [InlineKeyboardButton("🏧 Withdraw", callback_data="withdraw")],
    [InlineKeyboardButton("❓ Help", callback_data="help")]
])

ADMIN_MENU = InlineKeyboardMarkup([
    [InlineKeyboardButton("📂 Upload Numbers (CSV/XLSX)", callback_data="admin_upload")],
    [InlineKeyboardButton("👥 View Users", callback_data="admin_view_users"),
     InlineKeyboardButton("🧾 Withdraw Requests", callback_data="admin_withdrawals")],
    [InlineKeyboardButton("📤 Broadcast", callback_data="admin_broadcast")],
    [InlineKeyboardButton("🔙 Back", callback_data="back_to_main")]
])

# in-memory waiting states
AWAITING_BKASH = {}       # tg_id -> amount
AWAITING_UPLOAD = set()   # admin tg_id awaiting upload
AWAITING_BROADCAST = {}   # admin tg_id -> True

async def start_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    text = ("Welcome! 🎉\n"
            "Earn money by receiving OTPs. Use one of our numbers to receive OTPs and earn Tk per OTP.\n")
    if user.id == ADMIN_ID:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("Open Admin Panel", callback_data="open_admin")],
            [InlineKeyboardButton("Open Main Menu", callback_data="open_main")]
        ]))
    else:
        await update.message.reply_text(text, reply_markup=MAIN_MENU)

async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    uid = query.from_user.id
    users = load_users()
    numbers = load_numbers()

    # navigation shortcuts
    if data == "open_admin":
        if uid == ADMIN_ID:
            await query.edit_message_text("Admin panel:", reply_markup=ADMIN_MENU)
        else:
            await query.edit_message_text("You are not admin.", reply_markup=MAIN_MENU)
        return
    if data == "open_main":
        await query.edit_message_text("Menu:", reply_markup=MAIN_MENU)
        return

    # user actions
    if data == "get_number":
        u = users.get(str(uid))
        if u and u.get("number"):
            await query.edit_message_text(f"You already have an assigned number: {u['number']}", reply_markup=MAIN_MENU)
            return
        free = find_free_number(numbers)
        if not free:
            await query.edit_message_text("Sorry — no free numbers available right now.", reply_markup=MAIN_MENU)
            return
        username = query.from_user.username or query.from_user.first_name or ""
        assign_number_to_user(numbers, users, free, uid, username)
        await query.edit_message_text(
            f"✅ Number assigned: {free['number']}\nUse it to register on other platforms. You will receive Tk {REWARD_PER_OTP} for each OTP received.",
            reply_markup=MAIN_MENU
        )
        await send_message(context, TEAM_CHAT_ID, f"📥 Number {free['number']} assigned to @{username} (id:{uid}).")
        await send_message(context, ADMIN_ID, f"📥 Number {free['number']} assigned to @{username} (id:{uid}).")
        return

    if data == "my_balance":
        rec = users.get(str(uid))
        if not rec:
            await query.edit_message_text("You don't have a number yet. Use Get Number to claim one.", reply_markup=MAIN_MENU)
            return
        bal = rec.get("balance", 0)
        await query.edit_message_text(f"Your number: {rec.get('number')}\nBalance: ৳{bal:.2f}", reply_markup=MAIN_MENU)
        return

    if data == "my_otps":
        await query.edit_message_text("Your incoming OTPs are delivered privately to your chat when received.", reply_markup=MAIN_MENU)
        return

    if data == "help":
        await query.edit_message_text("Help:\n• Use Get Number to receive a number.\n• Use the number on external sites to receive OTPs.\n• Each OTP gives you Tk 1 credited to your balance.\n• Minimum withdrawal is ৳{MIN_WITHDRAW}.", reply_markup=MAIN_MENU)
        return

    if data == "withdraw":
        users = load_users()
        rec = users.get(str(uid))
        if not rec or not rec.get("number"):
            await query.edit_message_text("You need an assigned number to earn. Use Get Number first.", reply_markup=MAIN_MENU)
            return
        bal = rec.get("balance", 0)
        if bal < MIN_WITHDRAW:
            await query.edit_message_text(f"❌ Minimum withdrawal is ৳{MIN_WITHDRAW}. You currently have ৳{bal:.2f}.", reply_markup=MAIN_MENU)
            return
        # ask for bKash number
        AWAITING_BKASH[uid] = MIN_WITHDRAW
        await query.edit_message_text(f"Please send your bKash number now (e.g., 017XXXXXXXX). You will request ৳{MIN_WITHDRAW}.", reply_markup=None)
        return

    # admin actions
    if str(uid) != str(ADMIN_ID):
        await query.edit_message_text("Unknown command or you are not admin.", reply_markup=MAIN_MENU)
        return

    if data == "admin_upload":
        AWAITING_UPLOAD.add(uid)
        await query.edit_message_text("Please send the CSV or XLSX file with one column 'number' or one number per line as a document.")
        return

    if data == "admin_view_users":
        users = load_users()
        if not users:
            await query.edit_message_text("No users registered yet.", reply_markup=ADMIN_MENU)
            return
        lines = []
        for tg, u in users.items():
            lines.append(f"{tg} @{u.get('username','')}\nNumber: {u.get('number')}\nBalance: ৳{u.get('balance',0):.2f}\n")
        text = "\n".join(lines)
        if len(text) > 3500:
            bio = BytesIO(text.encode("utf-8"))
            bio.name = "users.txt"
            await context.bot.send_document(chat_id=ADMIN_ID, document=InputFile(bio))
            await query.edit_message_text("Sent full users list as file.", reply_markup=ADMIN_MENU)
        else:
            await query.edit_message_text(text, reply_markup=ADMIN_MENU)
        return

    if data == "admin_withdrawals":
        withds = load_withdrawals()
        pend = [w for w in withds if w.get("status") == "pending"]
        if not pend:
            await query.edit_message_text("No pending withdrawal requests.", reply_markup=ADMIN_MENU)
            return
        # list each and provide action buttons
        for w in pend:
            txt = (f"User: @{w.get('username','')} ({w.get('tg_id')})\n"
                   f"Amount: ৳{w.get('amount')}\n"
                   f"bKash: {w.get('bkash')}\n"
                   f"Requested: {w.get('requested_at')}\nID: {w.get('id')}")
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Release Payment", callback_data=f"release_{w.get('id')}"),
                 InlineKeyboardButton("❌ Decline", callback_data=f"decline_{w.get('id')}")]
            ])
            await context.bot.send_message(chat_id=ADMIN_ID, text=txt, reply_markup=kb)
        await query.edit_message_text("Displayed pending withdrawals to admin.", reply_markup=ADMIN_MENU)
        return

    if data == "admin_broadcast":
        AWAITING_BROADCAST[uid] = True
        await query.edit_message_text("Send the broadcast message now. It will be sent to all users.", reply_markup=None)
        return

    if data == "back_to_main":
        await query.edit_message_text("Admin menu:", reply_markup=ADMIN_MENU)
        return

# ---------------------------
# File upload handler (admin)
# ---------------------------
async def upload_document_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # only admin
    if update.message.from_user.id != ADMIN_ID:
        return
    if update.message.document is None:
        return
    if update.message.from_user.id not in AWAITING_UPLOAD:
        return
    doc = update.message.document
    file = await context.bot.get_file(doc.file_id)
    bio = BytesIO()
    await file.download_to_memory(out=bio)
    bio.seek(0)
    fname = (doc.file_name or "").lower()
    parsed = []
    try:
        if fname.endswith(".csv") or fname.endswith(".txt"):
            text = bio.read().decode("utf-8", errors="ignore")
            for row in text.splitlines():
                row = row.strip()
                if not row: continue
                parsed.append(row)
        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            wb = load_workbook(filename=BytesIO(bio.read()), read_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if not r: continue
                first = r[0]
                if first is None: continue
                parsed.append(str(first).strip())
        else:
            await update.message.reply_text("Unsupported file type. Use CSV or XLSX.")
            AWAITING_UPLOAD.discard(update.message.from_user.id)
            return
    except Exception as e:
        await update.message.reply_text(f"Failed to parse file: {e}")
        AWAITING_UPLOAD.discard(update.message.from_user.id)
        return

    if not parsed:
        await update.message.reply_text("No numbers found in file.")
        AWAITING_UPLOAD.discard(update.message.from_user.id)
        return

    existing = load_numbers()
    exist_set = {normalize_num(n.get("number")) for n in existing if n.get("number")}
    added = 0
    for p in parsed:
        if normalize_num(p) not in exist_set:
            existing.append({"number": p, "assigned_to": None})
            exist_set.add(normalize_num(p))
            added += 1
    save_numbers(existing)
    AWAITING_UPLOAD.discard(update.message.from_user.id)
    await update.message.reply_text(f"Uploaded numbers. Added {added} new numbers. Total pool: {len(existing)}")

# ---------------------------
# Text messages handler (withdraw, admin broadcast)
# ---------------------------
async def text_message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.message.from_user.id
    text = (update.message.text or "").strip()
    users = load_users()

    # Admin broadcast
    if uid == ADMIN_ID and AWAITING_BROADCAST.get(uid):
        AWAITING_BROADCAST.pop(uid, None)
        all_users = load_users()
        for tg_id, u in all_users.items():
            try:
                await context.bot.send_message(chat_id=int(tg_id), text=text)
            except Exception:
                pass
        await update.message.reply_text("Broadcast sent to all users.")
        return

    # User sending bKash number after Withdraw prompt
    if uid in AWAITING_BKASH:
        amount = AWAITING_BKASH.pop(uid, None)
        bk = text
        # basic validation
        if not re.match(r"^\+?\d{10,15}$", bk) and not re.match(r"^01\d{9}$", bk):
            await update.message.reply_text("Invalid bKash number format. Example: 017XXXXXXXX")
            return
        users = load_users()
        rec = users.get(str(uid))
        if not rec:
            await update.message.reply_text("Account not found. Get a number first.")
            return
        bal = rec.get("balance", 0)
        if bal < amount:
            await update.message.reply_text(f"Insufficient balance. You have ৳{bal:.2f}")
            return
        # record withdrawal
        withds = load_withdrawals()
        wid = str(uuid.uuid4())[:8]
        record = {
            "id": wid,
            "tg_id": str(uid),
            "username": rec.get("username", ""),
            "bkash": bk,
            "amount": amount,
            "status": "pending",
            "requested_at": datetime.utcnow().isoformat() + "Z"
        }
        withds.append(record)
        save_withdrawals(withds)
        await update.message.reply_text(f"✅ Withdrawal request recorded for ৳{amount}. Request ID: {wid}")
        # notify admin & team (full visibility)
        msg = (f"💵 New withdrawal request\nUser: @{record['username']} ({record['tg_id']})\n"
               f"Amount: ৳{record['amount']}\nBkash: {record['bkash']}\nRequest ID: {record['id']}")
        await send_message(context, ADMIN_ID, msg)
        await send_message(context, TEAM_CHAT_ID, msg)
        return

    # Other text: ignore or reply
    # Optionally respond to /help text here

# ---------------------------
# Withdrawal action callbacks (release/decline)
# ---------------------------
async def withdraw_action_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    if uid != ADMIN_ID:
        await query.edit_message_text("Only admin can perform this action.")
        return
    data = query.data
    if not data:
        return
    if data.startswith("release_") or data.startswith("decline_"):
        action, wid = data.split("_", 1)
        withds = load_withdrawals()
        rec = next((w for w in withds if w.get("id") == wid), None)
        if not rec:
            await query.edit_message_text("Request not found or already handled.")
            return
        if action == "decline":
            rec["status"] = "declined"
            rec["handled_at"] = datetime.utcnow().isoformat() + "Z"
            save_withdrawals(withds)
            try:
                await context.bot.send_message(chat_id=int(rec["tg_id"]), text=f"❌ Your withdrawal request (ID {wid}) was declined by admin.")
            except Exception:
                pass
            await context.bot.send_message(chat_id=TEAM_CHAT_ID, text=f"Withdrawal {wid} declined by admin.")
            await query.edit_message_text(f"Declined request {wid}.")
            return
        # release
        if action == "release":
            if rec.get("status") != "pending":
                await query.edit_message_text("Request already handled.")
                return
            users = load_users()
            user = users.get(rec.get("tg_id"))
            if not user:
                await query.edit_message_text("User not found.")
                return
            cur_bal = user.get("balance", 0)
            amt = rec.get("amount", 0)
            if cur_bal < amt:
                await query.edit_message_text("User does not have enough balance to release.")
                return
            # deduct and mark paid
            user["balance"] = cur_bal - amt
            rec["status"] = "paid"
            rec["handled_at"] = datetime.utcnow().isoformat() + "Z"
            rec["paid_by"] = str(uid)
            save_users(users)
            save_withdrawals(withds)
            try:
                await context.bot.send_message(chat_id=int(rec["tg_id"]),
                                               text=f"✅ Your withdrawal of ৳{amt} (ID {wid}) has been marked as PAID to bKash {rec.get('bkash')}.")
            except Exception:
                pass
            await context.bot.send_message(chat_id=TEAM_CHAT_ID,
                                          text=f"✅ Withdrawal {wid} for @{rec.get('username')} ({rec.get('tg_id')}) marked PAID by admin.")
            await query.edit_message_text(f"Released payment for {wid}. User notified.")
            return

# ---------------------------
# IVASMS job: poll inbox, credit users & forward messages
# ---------------------------
async def ivasms_job(context: ContextTypes.DEFAULT_TYPE):
    # run blocking web login + fetch in thread
    def _work():
        session = create_logged_session()
        if not session:
            return []
        return fetch_messages_from_page(session)
    try:
        msgs = await asyncio.to_thread(_work)
    except Exception as e:
        print("IVASMS job error:", e)
        msgs = []

    if not msgs:
        return

    seen = load_seen()
    users = load_users()
    numbers = load_numbers()
    new_flag = False

    for m in msgs:
        if m["id"] in seen:
            continue
        seen.add(m["id"])
        sender_or_to = m.get("sender_or_to","")
        content = m.get("content","")

        # try to extract OTP & metadata
        otp = extract_otp(content)
        provider, country = detect_provider_and_country(content + " " + sender_or_to)
        # find user
        tg_id, user_rec, matched_number = find_target_for_sms(users, numbers, sender_or_to, content)

        # Build message (exact format requested)
        # Keep monetary formatting to 2 decimal places
        username_disp = user_rec.get("username","") if user_rec else ""
        earned_text = f"৳{REWARD_PER_OTP:.4f}" if REWARD_PER_OTP < 1 else f"৳{REWARD_PER_OTP:.2f}"
        total_balance = 0.0
        if tg_id and user_rec:
            # credit reward
            users.setdefault(str(tg_id), user_rec)
            users[str(tg_id)]["balance"] = users[str(tg_id)].get("balance",0) + REWARD_PER_OTP
            total_balance = users[str(tg_id)]["balance"]
            save_users(users)

        # Format message (Markdown)
        msg_lines = [
            "📱 **New OTP!** ✨",
            f"📞 **Number:** {matched_number or sender_or_to}",
            f"🌍 **Country:** {country or 'Unknown'}",
            f"🆔 **Provider:** {provider or 'Unknown'}",
            f"🔑 **OTP Code:** {otp or 'N/A'}",
            f"📝 **Full Message:** {content}",
            "",
        ]
        if tg_id and user_rec:
            # user's name for display
            display_name = user_rec.get("username") or user_rec.get("username") or ""
            msg_lines.append(f"{display_name} **🎉 You have earned {earned_text} for this message!**")
            msg_lines.append(f"💰 **Total Balance:** ৳{total_balance:.2f}")
        else:
            msg_lines.append("No assigned user for this number.")
        final_msg = "\n".join(msg_lines)

        # send to team
        await send_message(context, TEAM_CHAT_ID, final_msg, parse_mode="Markdown")
        # send to admin as well
        await send_message(context, ADMIN_ID, final_msg, parse_mode="Markdown")

        # send to user (DM) privately
        if tg_id and user_rec:
            try:
                await send_message(context, int(tg_id), final_msg, parse_mode="Markdown")
            except Exception as e:
                print("Failed to send DM to user:", e)

        new_flag = True

    if new_flag:
        save_seen(seen)

# ---------------------------
# Application startup
# ---------------------------
def main():
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    # Handlers
    app.add_handler(CommandHandler("start", start_handler))
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, upload_document_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_message_handler))
    app.add_handler(CallbackQueryHandler(withdraw_action_callback, pattern=r'^(release_|decline_).+'))

    # schedule IVASMS job
    app.job_queue.run_repeating(ivasms_job, interval=POLL_SECONDS, first=10)

    print("Bot starting ...")
    app.run_polling()

if __name__ == "__main__":
    main()